from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.template.loader import get_template
from datetime import datetime
from django.db.models.functions import TruncMonth, ExtractWeekDay

from django.contrib.auth import authenticate, login
from django.db.models import Avg
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from django.urls import reverse
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.tokens import default_token_generator

from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from rest_framework.authtoken.models import Token
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, Avg, Sum, F
from core.models import Usuario, Estudiante, Admin, Profesor, Clase, Materia, Sesion, Evaluacion, ClaseMateria

from transbank.webpay.webpay_plus.transaction import Transaction
from transbank.common.options import WebpayOptions
from transbank.common.integration_type import IntegrationType
from transbank.common.integration_commerce_codes import IntegrationCommerceCodes
from transbank.common.integration_api_keys import IntegrationApiKeys
# Create your views here.


from django.shortcuts import render


def PaginaPrincipal(request):
    profesores = Profesor.objects.select_related('usuario').prefetch_related('clases_profesor').all()
    materias = Materia.objects.all()

    contexto = {
        "profesores": profesores,
        "user": request.user,
        "materias": materias
    }
    return render(request, 'core/html/PaginaPrincipal.html', contexto)
def ClasesLenguaje(request):
    # Obtener la instancia de la materia 'Lenguaje'
    materia_lenguaje = Materia.objects.get(id_materia=1)

    # Filtrar clases que tienen la materia 'Lenguaje'
    clases = Clase.objects.filter(materias=materia_lenguaje)

    return render(request, 'core/html/ClasesLenguaje.html', {'clases': clases})


def ClasesMatematica(request):
    # Obtener la instancia de la materia 'Matemática'
    materia_matematica = Materia.objects.get(id_materia=2)

    # Filtrar clases que tienen la materia 'Matemática'
    clases = Clase.objects.filter(materias=materia_matematica)

    return render(request, 'core/html/ClasesMatematica.html', {'clases': clases})


def ClasesHistoria(request):
    # Obtener la instancia de la materia 'Historia'
    materia_historia = Materia.objects.get(id_materia=3)

    # Filtrar clases que tienen la materia 'Historia'
    clases = Clase.objects.filter(materias=materia_historia)

    return render(request, 'core/html/ClasesHistoria.html', {'clases': clases})

def Clases(request):
    # Obtener la instancia de la materia 'Historia'


    # Filtrar clases que tienen la materia 'Historia'
    clases = Clase.objects.all()

    return render(request, 'core/html/Clases.html', {'clases': clases})




def Login(request):
    return render(request, 'core/Logueo/Login.html')


def Logueo(request):
    if request.method == 'POST':
        correo = request.POST.get('email1')
        contra = request.POST.get('contra1')

        # Buscar al usuario en el modelo personalizado
        try:
            usuario1 = Usuario.objects.get(email=correo)
        except Usuario.DoesNotExist:
            messages.error(request, 'No se encontró el usuario')
            return redirect('Login')

        # Validar la contraseña ingresada con la contraseña almacenada
        if contra == usuario1.contra:
            # Verificar el tipo de usuario
            if usuario1.tipo_de_usuario == "Profesor":
                try:
                    profesor = Profesor.objects.get(usuario=usuario1)
                    if profesor.estado_de_aprobacion == 'Pendiente':
                        messages.error(request, 'Tu cuenta está pendiente de aprobación. Por favor, espera la aprobación para iniciar sesión.')
                        return redirect('Login')
                except Profesor.DoesNotExist:
                    messages.error(request, 'Tu cuenta de profesor no está configurada correctamente.')
                    return redirect('Login')

            elif usuario1.tipo_de_usuario == "Estudiante":
                try:
                    estudiante = Estudiante.objects.get(usuario=usuario1)
                    if estudiante.estado_solicitud == 'Pendiente':
                        messages.error(request, 'Tu cuenta está pendiente de aprobación por parte de tus padres. Por favor, espera la aprobación para iniciar sesión.')
                        return redirect('Login')
                except Estudiante.DoesNotExist:
                    messages.error(request, 'Tu cuenta de estudiante no está configurada correctamente.')
                    return redirect('Login')

            # Autenticar al usuario utilizando authenticate
            user = authenticate(request, username=usuario1.email, password=usuario1.contra)
            if user is not None:
                login(request, user)
                token, created = Token.objects.get_or_create(user=user)
                # Redirigir según el tipo de usuario
                if usuario1.tipo_de_usuario == "Admin":
                    return redirect('PanelAdmin')
                elif usuario1.tipo_de_usuario == "Estudiante" or usuario1.tipo_de_usuario == "Profesor":
                    return redirect('Perfil')
            else:
                messages.error(request, 'La contraseña es incorrecta')
                return redirect('Login')
        else:
            messages.error(request, 'La contraseña es incorrecta')
            return redirect('Login')

    return redirect('Login')

def Deslogueo(request):
    logout(request)
    return redirect('PaginaPrincipal')

def Agendar (request, id_profesor):
    profe = Profesor.objects.select_related('usuario').get(id_profesor=id_profesor)
    usuario = Usuario.objects.get(email = request.user.username)
    clase = Clase.objects.get(profesor = profe)
    contexto = {
        "profe": profe,
        "clase": clase,
        "usuario": usuario
    }
    return render(request, 'core/html/Agendar.html', contexto)
    
def CambiarContra (request):
    return render(request, 'core/html/CambiarContra.html')

def Solicitudes(request):
    solicitudes = Profesor.objects.filter(estado_de_aprobacion="Pendiente")  # Solo las solicitudes pendientes
    return render(request, 'core/html/Solicitudes.html', {'solicitudes': solicitudes})

# Vista para aceptar una solicitud
def AceptarSolicitud(request, id_solicitud):
    try:
        profesor = Profesor.objects.get(id_profesor=id_solicitud)
        profesor.estado_de_aprobacion = "Aprobado"
        profesor.save()
        send_email(profesor.usuario.email, request, 'aprobado')
        messages.success(request, "Solicitud aceptada con éxito.")
    except Profesor.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    return redirect('Solicitudes')

# Vista para rechazar una solicitud
def RechazarSolicitud(request, id_solicitud):
    try:
        profesor = Profesor.objects.get(id_profesor=id_solicitud)
        usuario = profesor.usuario

        # Eliminar el usuario de Django asociado
        try:
            user = User.objects.get(username=usuario.email)
            user.delete()
        except User.DoesNotExist:
            pass  # Si no existe el usuario en la tabla de User, no hacemos nada

        # Eliminar el profesor y el usuario de la tabla personalizada
        profesor.delete()
        usuario.delete()

        send_email(usuario.email, request, 'rechazado')
        messages.success(request, "Solicitud rechazada y usuario eliminado con éxito.")
    except Profesor.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
    return redirect('Solicitudes')

def DetalleSolicitud(request, id_solicitud):
    profesor = get_object_or_404(Profesor, id_profesor=id_solicitud)
    return render(request, 'core/html/DetalleSolicitud.html', {'profesor': profesor})
    
def PanelAdmin(request):
    # Total de usuarios
    total_usuarios = Usuario.objects.count()
    total_profesores = Profesor.objects.count()
    total_solicitudes = Profesor.objects.filter(estado_de_aprobacion='Pendiente').count()

    # Distribución de usuarios por rol
    usuarios_por_rol = Usuario.objects.values('tipo_de_usuario').annotate(count=Count('tipo_de_usuario'))

    # Número total de clases
    total_clases = Clase.objects.count()

    # Clases por categoría/tema
    clases_por_categoria = Clase.objects.values('descripcion_clase').annotate(count=Count('descripcion_clase'))

    # Número de clases por profesor
    clases_por_profesor = Clase.objects.values('profesor__usuario__nombre').annotate(count=Count('id_clase'))

    # Evaluaciones promedio de las clases
    evaluacion_promedio = 0
    if Evaluacion.objects.exists():
        evaluacion_promedio = Evaluacion.objects.aggregate(avg_evaluacion=Avg('valoracion'))['avg_evaluacion']

    # Preferencias de los estudiantes por rangos de edad
    preferencias_edad = Sesion.objects.values('estudiante__usuario__edad').annotate(
        count_clases=Count('id_sesion'),
        avg_prof_edad=Avg('profesor__usuario__edad'),
        count_prof_genero=Count('profesor__usuario__sexo'),
    ).order_by('estudiante__usuario__edad')

    # Filtrar para obtener las edades y géneros más seleccionados
    for pref in preferencias_edad:
        est_edad = pref['estudiante__usuario__edad']
        edad_pref_profesor = Sesion.objects.filter(estudiante__usuario__edad=est_edad).values('profesor__usuario__edad').annotate(count=Count('profesor__usuario__edad')).order_by('-count')
        genero_pref_profesor = Sesion.objects.filter(estudiante__usuario__edad=est_edad).values('profesor__usuario__sexo').annotate(count=Count('profesor__usuario__sexo')).order_by('-count')

        pref['pref_prof_edad'] = [edad['profesor__usuario__edad'] for edad in edad_pref_profesor]
        
        # Obtener el género más seleccionado por los estudiantes y convertir a palabras
        if genero_pref_profesor.exists():
            genero_mas_seleccionado = genero_pref_profesor[0]['profesor__usuario__sexo']
            if genero_mas_seleccionado == 'Masculino':
                pref['pref_prof_genero'] = 'Masculino'
            elif genero_mas_seleccionado == 'Femenino':
                pref['pref_prof_genero'] = 'Femenino'
            else:
                pref['pref_prof_genero'] = 'Otros'
        else:
            pref['pref_prof_genero'] = 'No especificado'

        # Obtener las materias más seleccionadas por los estudiantes
        pref['pref_materia'] = []

        # Obtener los idiomas más seleccionados por los estudiantes
        pref['pref_idioma'] = []

    # Total de sesiones de clases completadas, canceladas y pendientes
    sesiones_por_estado = Sesion.objects.values('estado_clase').annotate(count=Count('estado_clase'))

    # Distribución de sesiones por mes y día de la semana
    sesiones_por_mes = Sesion.objects.annotate(month=TruncMonth('fechaclase')).values('month').annotate(count=Count('id_sesion'))
    sesiones_por_dia_semana = Sesion.objects.annotate(day_of_week=ExtractWeekDay('fechaclase')).values('day_of_week').annotate(count=Count('id_sesion'))

    # Usuarios más activos
    estudiantes_actividades = Estudiante.objects.annotate(count_sesiones=Count('sesion')).order_by('-count_sesiones')[:10]
    profesores_actividades = Profesor.objects.annotate(count_sesiones=Count('sesiones_profesor')).order_by('-count_sesiones')[:10]

    # Datos financieros
    ingresos_totales = Clase.objects.aggregate(total=Sum('tarifa_clase'))

    context = {
        'total_usuarios': total_usuarios,
        'total_profesores': total_profesores,
        'total_solicitudes': total_solicitudes,
        'usuarios_por_rol': usuarios_por_rol,
        'total_clases': total_clases,
        'clases_por_categoria': clases_por_categoria,
        'clases_por_profesor': clases_por_profesor,
        'evaluacion_promedio': evaluacion_promedio,
        'preferencias_edad': preferencias_edad,
        'sesiones_por_estado': sesiones_por_estado,
        'sesiones_por_mes': sesiones_por_mes,
        'sesiones_por_dia_semana': sesiones_por_dia_semana,
        'estudiantes_actividades': estudiantes_actividades,
        'profesores_actividades': profesores_actividades,
        'ingresos_totales': ingresos_totales,
    }

    return render(request, 'core/html/PanelAdmin.html', context)
    
# Vista para exportar datos a Excel
def exportar_excel(request):
    # Obtén los datos necesarios para exportar
    sesiones_por_estado = Clase.objects.values('estado_clase').annotate(count=Count('estado_clase'))

    # Crear un DataFrame con los datos
    df = pd.DataFrame({
        'Estado de Clase': [sesion['estado_clase'] for sesion in sesiones_por_estado],
        'Número de Sesiones': [sesion['count'] for sesion in sesiones_por_estado]
    })

    # Crear un archivo Excel y escribir los datos en él
    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos de Sesiones'

    # Escribir los datos del DataFrame en la hoja de trabajo
    for r_idx, row in enumerate(df.itertuples(), start=1):
        ws.cell(row=r_idx + 1, column=1, value=row[1])
        ws.cell(row=r_idx + 1, column=2, value=row[2])

    # Agregar un gráfico de barras
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Número de Sesiones por Estado de Clase"
    chart.x_axis.title = 'Estado de Clase'
    chart.y_axis.title = 'Número de Sesiones'

    data = Reference(ws, min_col=2, min_row=1, max_row=len(df) + 1, max_col=2)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    ws.add_chart(chart, "D10")

    # Guardar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sesiones_clases.xlsx"'
    wb.save(response)

    return response    
    
def PerfilProfe (request):
    return render(request, 'core/html/PerfilProfe.html')



def VistaProfe(request, id_profesor, id_clase):
    # Obtener el profesor y la clase correspondientes
    profe = Profesor.objects.select_related('usuario').get(id_profesor=id_profesor)
    clase = Clase.objects.select_related('profesor').get(id_clase=id_clase)
    cantResenas = Evaluacion.objects.filter(clase_id=id_clase).count()
    avgResena = Evaluacion.objects.filter(clase_id=id_clase).aggregate(promedio=Avg('valoracion'))['promedio']
    
    # Obtener todas las evaluaciones de la clase
    evaluaciones = Evaluacion.objects.filter(clase=clase)

    # Obtener el usuario actual si está autenticado
    usuario_actual = request.user
    estudiante = None
    evaluacion_existente = None

    if usuario_actual.is_authenticated:
        # Si el usuario está autenticado, intenta obtener el estudiante correspondiente y la evaluación existente
        try:
            estudiante = get_object_or_404(Estudiante, usuario__email=usuario_actual.email)
            evaluacion_existente = Evaluacion.objects.filter(
                profesor_id=id_profesor,
                estudiante=estudiante,
                clase_id=id_clase
            ).first()
        except Estudiante.DoesNotExist:
            # Si el usuario autenticado no es un estudiante, establecer evaluacion_existente como None
            pass
    
    # Pasar los datos al contexto
    contexto = {
        "profe": profe,
        "clase": clase,
        "evaluaciones": evaluaciones,
        "evaluacion_existente": evaluacion_existente,
        "cantResenas": cantResenas,
        "Usuario": usuario_actual,
        "avgResena":avgResena 
    }
    
    # Renderizar la plantilla con el contexto
    return render(request, 'core/html/VistaProfe.html', contexto)

def RegistroProfe(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        edad = request.POST.get('edad')
        sexo = request.POST.get('sexo')
        telefono = request.POST.get('telefono')
        descripcion = request.POST.get('descripcion')
        run = request.POST.get('run')
        foto = request.FILES.get('foto_profe')
        carnet = request.FILES.get('carnet')
        certificado = request.FILES.get('certificado')
        antecedentes = request.FILES.get('antecedentes')
        contra = request.POST.get('contra')

        # Verificar si el correo electrónico ya está en uso
        if Usuario.objects.filter(email=email).exists():
            messages.warning(request, 'El correo ya está en uso')
            return redirect('RegistroProfe')

        # Crear el usuario personalizado
        usuario = Usuario.objects.create(
            email=email,
            nombre=nombre,
            edad=edad,
            sexo = sexo,
            apellido=apellido,
            telefono=telefono,
            contra=contra,
            foto=foto,
            tipo_de_usuario="Profesor"
        )

        # Crear el usuario de Django asociado
        user = User.objects.create_user(
            username=email,
            email=email,
            password=contra,
            first_name=nombre,
            last_name=apellido
        )

        # Crear el profesor asociado al usuario personalizado
        profesor = Profesor.objects.create(
            usuario=usuario,
            antecedentes=antecedentes,
            run=run,
            carnet=carnet,
            certificado = certificado,
            descripcion=descripcion,
            estado_de_aprobacion="Pendiente"
        )

        send_email(email, request, 'registro')  # Enviar el correo de notificación

        messages.success(request, "Registro completado con éxito. Tu cuenta está pendiente de aprobación.")
        return redirect('Login')  # Redirigir después de registrar con éxito

    return render(request, 'core/html/RegistroProfe.html', context={"materias": Materia.objects.all()})

def RegistroEstudiante(request):
    
    return render(request, 'core/html/RegistroEstudiante.html')

def FormularioEstudiante(request):
    if request.method == 'POST':
        vFoto = request.FILES.get('fotoAlumno')
        vNombre = request.POST.get('nombre')
        vSexo = request.POST.get('sexo')
        vApellido = request.POST.get('apellido')
        vTelefono = request.POST.get('telefono')
        vCorreo = request.POST.get('email')
        vEdad = request.POST.get('edad')
        vCorreoPadre = request.POST.get('correo_padres')
        vClave = request.POST.get('contrasena')
        vNvlEducativo = request.POST.get('NvlEducativo')

        # Verificar si el correo electrónico ya está en uso
        if Usuario.objects.filter(email=vCorreo).exists():
            messages.warning(request, 'El correo ya está en uso')
            return redirect('RegistroEstudiante')

        # Crear el usuario personalizado
        usuario = Usuario.objects.create(
            email=vCorreo,
            nombre=vNombre,
            sexo=vSexo,
            apellido=vApellido,
            telefono=vTelefono,
            contra=vClave,
            edad = vEdad,
            foto=vFoto,
            tipo_de_usuario="Estudiante"
        )

        # Crear el usuario de Django asociado
        user = User.objects.create_user(
            username=vCorreo,
            email=vCorreo,
            password=vClave,
            first_name=vNombre,
            last_name=vApellido
        )

        if int(vEdad) > 18:
            estado_solicitud = "Aprobado"
        else:
            estado_solicitud = "Pendiente"

        # Crear el estudiante asociado al usuario personalizado
        estudiante = Estudiante.objects.create(
            usuario=usuario,
            correo_padre = vCorreoPadre,
            estado_solicitud = estado_solicitud,
            nivel_educativo=vNvlEducativo
        )

        messages.success(request, "Registro completado con éxito.")
        return redirect('Login')  # Redirigir después de registrar con éxito

def RegistroAdmin(request):
    

    return render(request, 'core/html/RegistroAdmin.html')

def FormularioAdmin(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        email = request.POST.get('email')
        contra = request.POST.get('contra')
        telefono = request.POST.get('telefono', '')
        foto = request.FILES.get('foto', None)  # Si hay una imagen, se procesa


        if Usuario.objects.filter(email=email).exists():
            messages.error(request, "Ya existe un usuario con este correo.")
            return render(request, 'core/html/RegistroAdmin.html')

        try:
            # Crear usuario personalizado
            usuario = Usuario(
                email=email,
                nombre=nombre,
                apellido=apellido,
                contra=contra,
                tipo_de_usuario='Admin',
                telefono=telefono,
                foto=foto,
            )
            usuario.save()

            # Crear usuario Django para autenticación
            user = User.objects.create_user(
                username=email,
                email=email,
                password=contra,
                first_name=nombre,
                last_name=apellido
            )

            # Crear el perfil de Admin
            admin = Admin(
                usuario=usuario,
                nombre=nombre,
            )
            admin.save()

            messages.success(request, "Administrador creado con éxito.")
            return redirect('PanelAdmin')  # Redirigir al panel del administrador

        except Exception as e:
            messages.error(request, f"Error al crear el administrador: {e}")

def CambiarContra(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            send_email(email, request, 'cambiar')  # Pasar el tipo 'cambiar'
            messages.success(request, 'Se ha enviado un correo para cambiar tu contraseña.')
            return redirect('CambiarContra')  # Redirigir a la misma página para mostrar el mensaje
        except User.DoesNotExist:
            messages.error(request, 'El correo electrónico no está registrado.')
    
    return render(request, 'core/html/CambiarContra.html')

def solicitar_cambio_contra(request, tipo):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            send_email(email, request, 'restablecer')
            messages.success(request, f'Se ha enviado un correo para {tipo} tu contraseña.')
            return redirect('Login')
        except User.DoesNotExist:
            messages.error(request, 'El correo electrónico no está registrado.')
    
    return render(request, 'core/html/solicitar_cambio_contra.html', {'tipo': tipo})

def send_email(email, request, tipo):
    context = {
        'email': email,
        'request': request,
        'tipo': tipo,
    }

    template = get_template('core/html/correo_generico.html')
    subject = 'Cambio de Contraseña' if tipo == 'cambiar' else 'Restablecimiento de Contraseña' if tipo == 'restablecer' else 'Estado de Registro'
    content = template.render(context)

    mail = EmailMultiAlternatives(
        subject,
        'Correo de notificación',
        settings.EMAIL_HOST_USER,
        [email]
    )

    mail.attach_alternative(content, 'text/html')
    mail.send()

def reset_password(request, email):
    if request.method == 'POST': 
        new_password = request.POST.get('password')
        
        try:
            user = User.objects.get(email=email)
            usuario = get_object_or_404(Usuario, email=email)
            
            # Actualizar la contraseña en ambas tablas
            user.set_password(new_password)
            user.save()
            
            usuario.contra = new_password
            usuario.save()

            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            return redirect('Login')
        except User.DoesNotExist:
            messages.error(request, 'El correo electrónico no está registrado.')
    
    return render(request, 'core/html/reset_password.html', {'email': email})
    
@login_required
def Perfil(request):
    usuario = Usuario.objects.get(email=request.user.username)
    try:
        profe = Profesor.objects.get(usuario=usuario)
    except Profesor.DoesNotExist:
        profe = None
    
    return render(request, 'core/html/Perfil.html', {'usuario': usuario, 'profe': profe})

def ModificarPerfil(request):
    usuario = request.user
    usuarioActivo = Usuario.objects.get(email=usuario.email)
    profe = Profesor.objects.get(usuario=usuarioActivo)

    if request.method == 'POST':
        vFoto = request.FILES.get('fotoPerfil')
        vNombre = request.POST.get('nombre')
        vApellido = request.POST.get('apellido')
        vTelefono = request.POST.get('telefono')
        vEdad = request.POST.get('edad')
        vDescripcion = request.POST.get('descripcion')

        try:
            # Actualizar campos del Usuario existente
            usuarioActivo.nombre = vNombre
            usuarioActivo.apellido = vApellido
            usuarioActivo.telefono = vTelefono
            usuarioActivo.edad = vEdad
            if vFoto:
                usuarioActivo.foto = vFoto
            usuarioActivo.save()

            # Actualizar campos del Profesor
            profe.descripcion = vDescripcion
            profe.save()

            print('Perfil actualizado exitosamente')
        except :
             print('Error al actualizar perfil')

        return redirect('Perfil')  # Ajusta 'Perfil' al nombre de tu vista de perfil


@login_required
def ListaUsuarios(request):
    usuarios = Usuario.objects.all()
    return render(request, 'core/html/ListaUsuarios.html', {'usuarios': usuarios})

@login_required
def ListaClases(request):
    clases = Sesion.objects.all()
    return render(request, 'core/html/ListaClases.html', {'clases': clases})

def EliminarUsuario(request, usuario_id):
    try:
        # Buscar el usuario por ID
        usuario = Usuario.objects.get(id_usuario=usuario_id)

        # Eliminar el usuario
        usuario.delete()

        # Mensaje de éxito
        messages.success(request, f"Usuario {usuario.nombre} {usuario.apellido} eliminado con éxito.")

    except Usuario.DoesNotExist:
        # Manejar caso donde el usuario no exista
        messages.error(request, "Usuario no encontrado.")

    except Exception as e:
        # Manejar otros errores inesperados
        messages.error(request, f"Error al eliminar usuario: {e}")

    # Redirigir a la lista de usuarios
    return redirect('ListaUsuarios')

def EliminarClase(request, clase_id):
    try:
        # Obtén la clase y elimínala
        clase = Sesion.objects.get(id_sesion=clase_id)
        clase.delete()
        messages.success(request, f"Clase eliminada con éxito.")
    except Sesion.DoesNotExist:
        messages.error(request, "Clase no encontrada.")
    except Exception as e:
        messages.error(request, f"Error al eliminar clase: {e}")

    # Redirigir a una página después de la eliminación, como la lista de clases
    return redirect('ListaClases')

def VerClase(request, clase_id):
    # Obtén la clase correspondiente al ID o devuelve un 404 si no existe
    clase = get_object_or_404(Sesion, id_sesion=clase_id)
    return render(request, 'core/html/VerClase.html', {'clase': clase})

def FormClase(request):
    if request.method == 'POST':
        # Obtener datos del formulario
        titulo = request.POST.get('titulo')
        descripcion = request.POST.get('descripcion')
        precio = request.POST.get('precio')
        materia = request.POST.get('materia')
        
        # Obtener el usuario autenticado (profesor)
        usuario = request.user
        
        try:
            # Obtener el objeto Usuario correspondiente al usuario autenticado
            print(materia)
            usuario1 = Usuario.objects.get(email=usuario)
            
            # Obtener el objeto Profesor correspondiente al usuario
            idusuario = Profesor.objects.get(usuario=usuario1)
            
            # Obtener el objeto Materia correspondiente al nombre de la materia
            materiaid = Materia.objects.get(id_materia=materia)
            
            # Crear la clase
            clase = Clase(
                nombre_clase=titulo,
                tarifa_clase=precio,
                descripcion_clase=descripcion,
                profesor=idusuario                
            )
            clase.save()
            
            # Relacionar la clase con la materia a través de la tabla intermedia ClaseMateria
            claseMateria = ClaseMateria(
                clase_id=clase.id_clase,
                materia_id=materiaid.id_materia
            )
            claseMateria.save()
            
            # Redireccionar al perfil del usuario después de guardar la clase
            return redirect('ClasesProfe')
            
        except Exception as e:
            messages.error(request, f"Error al crear la clase: {e}")
            return redirect('CrearClase')


def CrearClase(request):
    materias = Materia.objects.all()
    return render(request, 'core/html/FormClase.html', {'materias': materias })            

def ClasesProfe(request):
    usuario = request.user
    usuario1 = Usuario.objects.get(email = usuario)
    profe = Profesor.objects.get(usuario = usuario1)
    clases = Clase.objects.filter(profesor=profe).distinct()

    return render (request, 'core/html/VerClases.html', {'clases' : clases} )

def EditarClase(request, id_clase):
    if request.method == 'POST':
        clase = Clase.objects.get(pk=id_clase)
        
   
        clase.nombre_clase = request.POST.get('nombreClase')
        clase.tarifa_clase = request.POST.get('tarifaClase')
        clase.descripcion_clase = request.POST.get('descripcionClase')

    
        clase.save()

     
        return redirect('ClasesProfe')
    else:
       
        
        return render(request, 'PaginaPrincipal')

def EliminarClase(request,id_clase):
    clase = Clase.objects.get(id_clase = id_clase)
    clase.delete()
    return redirect('ClasesProfe')
    

def Agendar (request, id_profesor, id_clase):
    profe = Profesor.objects.select_related('usuario').get(id_profesor=id_profesor)
    usuario = Usuario.objects.get(email = request.user.username)
    estudiante = Estudiante.objects.get(usuario = usuario)
    clase = Clase.objects.select_related('profesor').get(id_clase=id_clase)
    sesion = Sesion.objects.get(estudiante = estudiante)
    contexto = {
        "profe": profe,
        "clase": clase,
        "usuario": usuario,
        'estudiante':estudiante,
        'sesion': sesion
    }
    return render(request, 'core/html/Agendar.html', contexto)

tx = Transaction(WebpayOptions(IntegrationCommerceCodes.WEBPAY_PLUS, IntegrationApiKeys.WEBPAY, IntegrationType.TEST))

def pagar(request, sesion_id):
    sesion = get_object_or_404(Sesion, id_sesion=sesion_id)

    # Datos de la transacción
    buy_order = str(sesion.id_sesion)
    session_id = str(sesion.estudiante.id_estudiante)
    amount = sesion.clase.tarifa_clase
    return_url = request.build_absolute_uri(reverse('retorno'))

    response = tx.create(buy_order, session_id, amount, return_url)

    token = response['token']
    url = response['url']

    return render(request, 'core/html/pagos/pago_formulario.html', {'url': url, 'token': token})

def retorno(request):
    token_ws = request.GET.get('token_ws', None)
    if token_ws:
        response = tx.commit(token_ws)
        if response['response_code'] == 0 and response['status'] == 'AUTHORIZED':
            # Transacción aprobada
            return render(request, 'core/html/pagos/pago_exitoso.html', {'response': response})
        else:
            # Transacción rechazada
            return render(request, 'core/html/pagos/pago_rechazado.html', {'response': response})
    else:
        return HttpResponse("Error: no se recibió token_ws")


def FormularioAgendar(request):
     if request.method == 'POST':
        mensaje = request.POST.get('mensaje')
        fecha_str = request.POST.get('datepicker')
        hora_str = request.POST.get('timepicker')
        telefono = request.POST.get('telefono')
        id_profesor = request.POST.get('id_profesor')
        usuario_actual = request.user
        usuario = get_object_or_404(Usuario, email=usuario_actual.email)
        estudiante = get_object_or_404(Estudiante, usuario=usuario)
        id_clase = request.POST.get('id_clase') 
        id_estudiante = estudiante.id_estudiante

        # Combinar fecha y hora
        datetime_str = f"{fecha_str} {hora_str}"
        fecha_hora = datetime.strptime(datetime_str, '%d/%m/%Y %H:%M')

        nueva_sesion = Sesion.objects.create(
            mensaje=mensaje,
            fechaclase=fecha_hora,
            contacto=telefono,
            profesor_id=id_profesor,
            estudiante_id=id_estudiante,
            clase_id=id_clase
        )
        print("ID del estudiante:", id_estudiante)

        return render (request, 'core/html/Agendar.html')


def Calificar(request, id_profesor, id_clase):
    # Obtener el usuario actual y el estudiante asociado
    usuario_actual = request.user
    estudiante = get_object_or_404(Estudiante, usuario__email=usuario_actual.email)

    # Intentar obtener la evaluación existente
    try:
        evaluacion_existente = Evaluacion.objects.get(
            profesor_id=id_profesor,
            estudiante=estudiante,
            clase_id=id_clase
        )
    except: evaluacion_existente = None

    if request.method == 'POST':
        # Obtener la calificación y el comentario del formulario
        calificacion = request.POST.get('calificacion')
        comentario = request.POST.get('comentario')

        # Si ya existe una evaluación, actualizarla; de lo contrario, crear una nueva
        if evaluacion_existente:
            evaluacion_existente.valoracion = calificacion
            evaluacion_existente.recomendacion = comentario
            evaluacion_existente.save()
        else:
            Evaluacion.objects.create(
                recomendacion=comentario,
                valoracion=calificacion,
                profesor_id=id_profesor,
                estudiante=estudiante,
                clase_id=id_clase
            )

        # Redireccionar a la vista del profesor después de la calificación
        return redirect('VistaProfe', id_profesor=id_profesor, id_clase=id_clase)


def ValidacionPapas(request,correo):
    Alumnos = Estudiante.objects.filter(estado_solicitud="Pendiente")  # Solo las solicitudes pendientes
    return render(request, 'core/html/ValidacionPapas.html', {'Alumnos': Alumnos})

def CorreoPapas(request):

    return render(request, 'core/html/CorreoPapas.html')

def ValidacionCorreoPapa(request):
    correo = request.POST.get('correo')
    try:
        estudiante = Estudiante.objects.get(correo_padre=correo)
        if estudiante:
            return redirect('ValidacionPapas', correo=correo)
    except Estudiante.DoesNotExist:
        messages.error(request, 'Correo no encontrado')
        return redirect('CorreoPapas')
    


def AceptarSolicitudEstudiante(request,id_estudiante ):
    try:
        estudiante = Estudiante.objects.get(id_estudiante=id_estudiante)
        estudiante.estado_solicitud = "Aprobado"
        estudiante.save()
        send_email(estudiante.usuario.email, request, 'aprobado')
        send_email(estudiante.correo_padre, request, 'Usted ha aprobado el registro de su hijo')
        messages.success(request, "Solicitud aceptada con éxito.")
    except Estudiante.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    return redirect('Login')

# Vista para rechazar una solicitud
def RechazarSolicitudEstudiante(request, id_estudiante):
    try:
        estudiante = Estudiante.objects.get(id_estudiante=id_estudiante)
        usuario = estudiante.usuario

        # Eliminar el usuario de Django asociado
        try:
            user = User.objects.get(username=usuario.email)
            user.delete()
        except User.DoesNotExist:
            pass  # Si no existe el usuario en la tabla de User, no hacemos nada

        # Eliminar el profesor y el usuario de la tabla personalizada
        estudiante.delete()
        usuario.delete()

        send_email(usuario.email, request, 'rechazado')
        messages.success(request, "Solicitud rechazada y usuario eliminado con éxito.")
    except Estudiante.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
    return redirect('Solicitudes')

   














    # def SolicitudesClase(request):
    #     solicitudes = Profesor.objects.filter(estado_clase="Pendiente")  # Solo las solicitudes pendientes
    # return render(request, 'core/html/VerClases.html', {'solicitudes': solicitudes})

    # def AceptarClase(request, id_solicitud):
    # try:
    #     profesor = Profesor.objects.get(id_profesor=id_solicitud)
    #     profesor.estado_de_aprobacion = "Aprobado"  # Cambiar el estado a Aprobado
    #     profesor.save()  # Guardar los cambios
    #     messages.success(request, "Solicitud aceptada con éxito.")
    # except Profesor.DoesNotExist:
    #     messages.error(request, "Solicitud no encontrada.")
    # return redirect('Solicitudes')  # Redirigir a la página de solicitudes
